from PyQt5 import QtCore, QtGui, QtWidgets

### START ###
from PyQt5.QtWidgets import QApplication, QWidget, QInputDialog, QLineEdit, QFileDialog
from PyQt5.QtGui import QIcon
import os
import pdfplumber
import re
from openpyxl import Workbook
from openpyxl import load_workbook
import time
main_file_list = {}
file_path = ""
### END ###


class Ui_MainWindow(object):
    def setupUi(self, MainWindow):
        MainWindow.setObjectName("MainWindow")
        MainWindow.resize(934, 583)
        self.centralwidget = QtWidgets.QWidget(MainWindow)
        self.centralwidget.setObjectName("centralwidget")
        self.listWidget = QtWidgets.QListWidget(self.centralwidget)
        self.listWidget.setGeometry(QtCore.QRect(340, 10, 581, 531))
        self.listWidget.setObjectName("listWidget")
        self.Title_lablel = QtWidgets.QLabel(self.centralwidget)
        self.Title_lablel.setGeometry(QtCore.QRect(20, 10, 271, 41))
        font = QtGui.QFont()
        font.setPointSize(16)
        font.setBold(True)
        font.setUnderline(True)
        font.setWeight(75)
        self.Title_lablel.setFont(font)
        self.Title_lablel.setObjectName("Title_lablel")
        self.label_enter_path = QtWidgets.QLabel(self.centralwidget)
        self.label_enter_path.setGeometry(QtCore.QRect(10, 60, 231, 21))
        self.label_enter_path.setObjectName("label_enter_path")
        self.pushButton_enter_path = QtWidgets.QPushButton(self.centralwidget)
        self.pushButton_enter_path.setGeometry(QtCore.QRect(216, 60, 120, 20))
        self.pushButton_enter_path.setObjectName("pushButton_enter_path")
        self.pushButton_rename = QtWidgets.QPushButton(self.centralwidget)
        self.pushButton_rename.setGeometry(QtCore.QRect(10, 490, 321, 51))
        font = QtGui.QFont()
        font.setPointSize(19)
        font.setBold(True)
        font.setWeight(75)
        self.pushButton_rename.setFont(font)
        self.pushButton_rename.setObjectName("pushButton_rename")
        self.lineEdit_enter_beginningtext = QtWidgets.QLineEdit(self.centralwidget)
        self.lineEdit_enter_beginningtext.setGeometry(QtCore.QRect(10, 140, 251, 20))
        self.lineEdit_enter_beginningtext.setObjectName("lineEdit_enter_beginningtext")
        self.pushButton_enter_beginningtext = QtWidgets.QPushButton(self.centralwidget)
        self.pushButton_enter_beginningtext.setGeometry(QtCore.QRect(263, 139, 75, 23))
        self.pushButton_enter_beginningtext.setObjectName("pushButton_enter_beginningtext")
        self.label_enter_beginningtext = QtWidgets.QLabel(self.centralwidget)
        self.label_enter_beginningtext.setGeometry(QtCore.QRect(10, 120, 231, 21))
        self.label_enter_beginningtext.setObjectName("label_enter_beginningtext")
        self.label_pathname = QtWidgets.QLabel(self.centralwidget)
        self.label_pathname.setGeometry(QtCore.QRect(10, 80, 311, 20))
        self.label_pathname.setObjectName("label_pathname")
        self.lineEdit_enter_endtext = QtWidgets.QLineEdit(self.centralwidget)
        self.lineEdit_enter_endtext.setGeometry(QtCore.QRect(11, 191, 250, 20))
        self.lineEdit_enter_endtext.setObjectName("lineEdit_enter_endtext")
        self.label_enter_endtext = QtWidgets.QLabel(self.centralwidget)
        self.label_enter_endtext.setGeometry(QtCore.QRect(11, 171, 230, 21))
        self.label_enter_endtext.setObjectName("label_enter_endtext")
        self.pushButton_enter_endtext = QtWidgets.QPushButton(self.centralwidget)
        self.pushButton_enter_endtext.setGeometry(QtCore.QRect(264, 190, 74, 23))
        self.pushButton_enter_endtext.setObjectName("pushButton_enter_endtext")
        self.lineEdit_enter_replace_text_before = QtWidgets.QLineEdit(self.centralwidget)
        self.lineEdit_enter_replace_text_before.setGeometry(QtCore.QRect(60, 239, 201, 20))
        self.lineEdit_enter_replace_text_before.setObjectName("lineEdit_enter_replace_text_before")
        self.pushButton_enter_replace_text = QtWidgets.QPushButton(self.centralwidget)
        self.pushButton_enter_replace_text.setGeometry(QtCore.QRect(264, 240, 74, 41))
        self.pushButton_enter_replace_text.setObjectName("pushButton_enter_replace_text")
        self.label_enter_replace_text = QtWidgets.QLabel(self.centralwidget)
        self.label_enter_replace_text.setGeometry(QtCore.QRect(11, 219, 230, 21))
        self.label_enter_replace_text.setObjectName("label_enter_replace_text")
        self.label_enter_replace_text_2 = QtWidgets.QLabel(self.centralwidget)
        self.label_enter_replace_text_2.setGeometry(QtCore.QRect(20, 237, 41, 21))
        self.label_enter_replace_text_2.setObjectName("label_enter_replace_text_2")
        self.lineEdit_enter_replace_text_after = QtWidgets.QLineEdit(self.centralwidget)
        self.lineEdit_enter_replace_text_after.setGeometry(QtCore.QRect(60, 262, 201, 20))
        self.lineEdit_enter_replace_text_after.setObjectName("lineEdit_enter_replace_text_after")
        self.label_enter_replace_text_3 = QtWidgets.QLabel(self.centralwidget)
        self.label_enter_replace_text_3.setGeometry(QtCore.QRect(20, 260, 41, 21))
        self.label_enter_replace_text_3.setObjectName("label_enter_replace_text_3")
        self.Title_lablel_2 = QtWidgets.QLabel(self.centralwidget)
        self.Title_lablel_2.setGeometry(QtCore.QRect(140, 284, 41, 41))
        font = QtGui.QFont()
        font.setPointSize(18)
        font.setBold(True)
        font.setUnderline(True)
        font.setWeight(75)
        self.Title_lablel_2.setFont(font)
        self.Title_lablel_2.setObjectName("Title_lablel_2")
        self.Title_lablel_3 = QtWidgets.QLabel(self.centralwidget)
        self.Title_lablel_3.setGeometry(QtCore.QRect(140, 359, 41, 41))
        font = QtGui.QFont()
        font.setPointSize(18)
        font.setBold(True)
        font.setUnderline(True)
        font.setWeight(75)
        self.Title_lablel_3.setFont(font)
        self.Title_lablel_3.setObjectName("Title_lablel_3")
        self.lineEdit_enter_titleblock_search = QtWidgets.QLineEdit(self.centralwidget)
        self.lineEdit_enter_titleblock_search.setGeometry(QtCore.QRect(10, 341, 250, 20))
        self.lineEdit_enter_titleblock_search.setObjectName("lineEdit_enter_titleblock_search")
        self.pushButton_enter_titleblock_search = QtWidgets.QPushButton(self.centralwidget)
        self.pushButton_enter_titleblock_search.setGeometry(QtCore.QRect(263, 340, 74, 23))
        self.pushButton_enter_titleblock_search.setObjectName("pushButton_enter_titleblock_search")
        self.label_enter_titleblock_search = QtWidgets.QLabel(self.centralwidget)
        self.label_enter_titleblock_search.setGeometry(QtCore.QRect(10, 321, 311, 21))
        self.label_enter_titleblock_search.setObjectName("label_enter_titleblock_search")
        self.label_enter_endtext_3 = QtWidgets.QLabel(self.centralwidget)
        self.label_enter_endtext_3.setGeometry(QtCore.QRect(10, 390, 331, 31))
        self.label_enter_endtext_3.setWordWrap(True)
        self.label_enter_endtext_3.setObjectName("label_enter_endtext_3")
        self.pushButton_enter_excel_export = QtWidgets.QPushButton(self.centralwidget)
        self.pushButton_enter_excel_export.setGeometry(QtCore.QRect(10, 420, 150, 41))
        self.pushButton_enter_excel_export.setObjectName("pushButton_enter_excel_export")
        self.line = QtWidgets.QFrame(self.centralwidget)
        self.line.setGeometry(QtCore.QRect(17, 470, 311, 20))
        self.line.setFrameShape(QtWidgets.QFrame.HLine)
        self.line.setFrameShadow(QtWidgets.QFrame.Sunken)
        self.line.setObjectName("line")
        self.pushButton_enter_excel_export_2 = QtWidgets.QPushButton(self.centralwidget)
        self.pushButton_enter_excel_export_2.setGeometry(QtCore.QRect(180, 420, 150, 41))
        self.pushButton_enter_excel_export_2.setObjectName("pushButton_enter_excel_export_2")
        MainWindow.setCentralWidget(self.centralwidget)
        self.menubar = QtWidgets.QMenuBar(MainWindow)
        self.menubar.setGeometry(QtCore.QRect(0, 0, 934, 21))
        self.menubar.setObjectName("menubar")
        self.menuAbout = QtWidgets.QMenu(self.menubar)
        self.menuAbout.setObjectName("menuAbout")
        MainWindow.setMenuBar(self.menubar)
        self.statusbar = QtWidgets.QStatusBar(MainWindow)
        self.statusbar.setObjectName("statusbar")
        MainWindow.setStatusBar(self.statusbar)
        self.actionAuthor = QtWidgets.QAction(MainWindow)
        self.actionAuthor.setObjectName("actionAuthor")
        self.menuAbout.addAction(self.actionAuthor)
        self.menubar.addAction(self.menuAbout.menuAction())

        ### START ###
        self.pushButton_enter_path.clicked.connect(self.enter_path)
        self.pushButton_enter_beginningtext.clicked.connect(self.enter_beginningtext)
        self.pushButton_enter_endtext.clicked.connect(self.enter_endtext)
        self.pushButton_enter_replace_text.clicked.connect(self.enter_replace_text)
        self.pushButton_enter_titleblock_search.clicked.connect(self.enter_titleblock_search)
        self.pushButton_rename.clicked.connect(self.rename_file)
        self.pushButton_enter_excel_export.clicked.connect(self.excel_list_export)
        self.pushButton_enter_excel_export_2.clicked.connect(self.excel_list_import)
        self.actionAuthor.triggered.connect(self.abouttheauthor)
        ### END ###

        self.retranslateUi(MainWindow)
        QtCore.QMetaObject.connectSlotsByName(MainWindow)

    def retranslateUi(self, MainWindow):
        _translate = QtCore.QCoreApplication.translate
        MainWindow.setWindowTitle(_translate("MainWindow", "MainWindow"))
        self.Title_lablel.setText(_translate("MainWindow", "Craig's Drawing Renamer"))
        self.label_enter_path.setToolTip(_translate("MainWindow",
                                                    "<html><head/><body><p>e.g. \'C:/Users/Craig/Desktop/Foldername\'</p></body></html>"))
        self.label_enter_path.setText(_translate("MainWindow", "Open the folder with the files to rename:"))
        self.pushButton_enter_path.setText(_translate("MainWindow", "Open"))
        self.pushButton_rename.setText(_translate("MainWindow", "Rename"))
        self.pushButton_enter_beginningtext.setText(_translate("MainWindow", "Preview"))
        self.label_enter_beginningtext.setToolTip(_translate("MainWindow",
                                                             "<html><head/><body><p>e.g. &quot;Your_Filename&quot; becomes &quot;Additional Text Your Filename&quot;</p></body></html>"))
        self.label_enter_beginningtext.setText(
            _translate("MainWindow", "Enter text to add to the start of the filename:"))
        self.label_pathname.setText(_translate("MainWindow", "<<Filepath>>"))
        self.label_enter_endtext.setToolTip(_translate("MainWindow",
                                                       "<html><head/><body><p>e.g. &quot;Your_Filename&quot; becomes &quot;Your Filename Additional Text&quot;</p></body></html>"))
        self.label_enter_endtext.setText(_translate("MainWindow", "Enter text to add to the end of the filename:"))
        self.pushButton_enter_endtext.setText(_translate("MainWindow", "Preview"))
        self.pushButton_enter_replace_text.setText(_translate("MainWindow", "Preview"))
        self.label_enter_replace_text.setToolTip(_translate("MainWindow",
                                                            "<html><head/><body><p>Leave the \'after\' box blank if you want text deleted from the filename</p></body></html>"))
        self.label_enter_replace_text.setText(_translate("MainWindow", "Replace text within a filename:"))
        self.label_enter_replace_text_2.setToolTip(_translate("MainWindow",
                                                              "<html><head/><body><p>e.g. &quot;Your_Filename&quot; becomes &quot;Additional Text Your Filename&quot;</p></body></html>"))
        self.label_enter_replace_text_2.setText(_translate("MainWindow", "Before"))
        self.label_enter_replace_text_3.setToolTip(_translate("MainWindow",
                                                              "<html><head/><body><p>e.g. &quot;Your_Filename&quot; becomes &quot;Additional Text Your Filename&quot;</p></body></html>"))
        self.label_enter_replace_text_3.setText(_translate("MainWindow", "After"))
        self.Title_lablel_2.setText(_translate("MainWindow", "Or"))
        self.Title_lablel_3.setText(_translate("MainWindow", "Or"))
        self.lineEdit_enter_titleblock_search.setText(_translate("MainWindow", "<<Enter Sample Drawing Number>>"))
        self.pushButton_enter_titleblock_search.setText(_translate("MainWindow", "Preview"))
        self.label_enter_titleblock_search.setToolTip(_translate("MainWindow",
                                                                 "<html><head/><body><p>Enter a sample drawing number as the programme searches for the same format</p></body></html>"))
        self.label_enter_titleblock_search.setText(
            _translate("MainWindow", "Attempt to rename the file automatically from the title block:"))
        self.label_enter_endtext_3.setToolTip(_translate("MainWindow",
                                                         "<html><head/><body><p>Enter a sample drawing number as the programme searches for the same format</p></body></html>"))
        self.label_enter_endtext_3.setText(_translate("MainWindow", "Use Excel to rename the files:"))
        self.pushButton_enter_excel_export.setText(_translate("MainWindow", "1.Export"))
        self.pushButton_enter_excel_export_2.setText(_translate("MainWindow", "2.Import"))
        self.menuAbout.setTitle(_translate("MainWindow", "About"))
        self.actionAuthor.setText(_translate("MainWindow", "Author"))

    ### START ###
    def enter_path(self):
        global file_path
        main_file_list.clear()
        file_path = QFileDialog.getExistingDirectory(None, "Open a folder", "C:\\", QFileDialog.ShowDirsOnly)
        self.label_pathname.setText(file_path)
        self.listWidget.clear()
        for file_name in os.listdir(file_path):
            main_file_list[file_name] = file_name
        for file_name in main_file_list:
            self.listWidget.addItem(file_name)
        return

    def enter_beginningtext(self):
        beginning_text = self.lineEdit_enter_beginningtext.text()
        self.listWidget.clear()
        for item in main_file_list:
            current_text = main_file_list[item]
            update_text = f'{beginning_text}{current_text}'
            main_file_list[item] = update_text
        for key, values in main_file_list.items():
            text_to_print = f'{key} =====> {values}'
            self.listWidget.addItem(text_to_print)
        return

    def enter_endtext(self):
        end_text = self.lineEdit_enter_endtext.text()
        self.listWidget.clear()
        for item in main_file_list:
            file_extension = item.split(".")[-1]
            current_text = main_file_list[item].split(".")[0]
            update_text = f'{current_text}{end_text}.{file_extension}'
            main_file_list[item] = update_text
        for key, values in main_file_list.items():
            text_to_print = f'{key} =====> {values}'
            self.listWidget.addItem(text_to_print)
        return

    def enter_replace_text(self):
        replace_text_before = self.lineEdit_enter_replace_text_before.text()
        replace_text_after = self.lineEdit_enter_replace_text_after.text()
        self.listWidget.clear()
        for item in main_file_list:
            current_text = main_file_list[item]
            update_text = current_text.replace(replace_text_before, replace_text_after)
            main_file_list[item] = update_text
            #print(main_file_list[item])
        for key, values in main_file_list.items():
            text_to_print = f'{key} =====> {values}'
            self.listWidget.addItem(text_to_print)
        return

    def enter_titleblock_search(self):
        # rename the file based on what is in the titleblock
        global file_path
        sample_drawing_number = self.lineEdit_enter_titleblock_search.text()
        self.listWidget.clear()
        self.listWidget.addItem("Please wait - this may take a while...")
        self.listWidget.addItem("Do not close window until complete - new drawings will list as they are found")
        # open each file in the list and extract text
        for item in main_file_list:
            file_extension = item.split(".")[-1]
            full_file_path = f'{file_path}/{item}'
            try:
                # Open the file and extract text with PyPDF2
                # Open the file and extract text with PyPDF2
                with pdfplumber.open(full_file_path) as pdf:
                    full_text = ""
                    pages = pdf.pages
                    for i, pg in enumerate(pages):
                        text = pages[i].extract_text()
                        full_text += text
                    # Convert the sample filename into regex search and search the extracted text
                    output_text = ""
                    regex_variable = "."
                    for char in sample_drawing_number:
                        if char == "-":
                            output_text += "-"
                        elif char == "_":
                            output_text += "_"
                        elif char == "(":
                            output_text += "\("
                        elif char == ")":
                            output_text += "\)"
                        else:
                            output_text += regex_variable
                    regex_results = re.findall(output_text, full_text)
                    number_of_regex_groups = len(regex_results) - 1
                    best_guess_drawing_number = regex_results[number_of_regex_groups]
                # Add the best guess to drawing number
                main_file_list[item] = f'{best_guess_drawing_number}.{file_extension}'
                self.listWidget.addItem(best_guess_drawing_number)
                QtCore.QCoreApplication.processEvents()
            # If it all goes wrong leave filename as is
            except Exception as e:
                main_file_list[item] = item
                self.listWidget.addItem(e)
        self.listWidget.clear()
        self.listWidget.addItem("Search Complete:")
        for key, values in main_file_list.items():
            text_to_print = f'{key} =====> {values}'
            self.listWidget.addItem(text_to_print)
        return

    def rename_file(self):
        self.listWidget.clear()
        self.listWidget.addItem("Renaming files")
        for key, values in main_file_list.items():
            new_file_name = file_path + "/" + values
            old_file_name = file_path + "/" + key
            os.rename(old_file_name, new_file_name)
            self.listWidget.addItem("Renaming files")
            text_to_print = f'{old_file_name} =====> {new_file_name}'
            self.listWidget.addItem(text_to_print)

    def excel_list_export(self):
        try:
            drawing_list = Workbook()
            worksheet_1 = drawing_list.active
            drawing_list_file_path = file_path + "/" + "Drawing_List.xlsx"
            drawing_list.save(drawing_list_file_path)
            row_number = 1
            self.listWidget.clear()
            self.listWidget.addItem("Saving files to excel sheet...")
            for key, values in main_file_list.items():
                file_name_no_extension = ""
                for char in key:
                    if char != ".":
                        file_name_no_extension += char
                    else:
                        break
                file_extension = key.split(".")[-1]
                worksheet_1.cell(row=row_number, column=1).value = file_name_no_extension
                worksheet_1.cell(row=row_number, column=2).value = file_extension
                row_number += 1
            drawing_list.save(drawing_list_file_path)
            self.listWidget.clear()
            self.listWidget.addItem("Excel file called \'Drawing_list\' has been saved in the same folder as the drawings")
            self.listWidget.addItem("\n\nAdd your new drawing numbers into Column C")
            self.listWidget.addItem("\n\nOnce complete make sure the file is closed, then press \'2.Import\'")
        except Exception as e:
            self.listWidget.clear()
            self.listWidget.addItem("Something has gone wrong - see exception below:\n\n")
            self.listWidget.addItem(e)

    def excel_list_import(self):
        try:
            drawing_list = Workbook()
            worksheet_1 = drawing_list.active
            drawing_list_file_path = file_path + "/" + "Drawing_List.xlsx"
            drawing_list = load_workbook(drawing_list_file_path)
            worksheet_1 = drawing_list.active
            row_number = worksheet_1.max_row
            row_number += 1 # to adjust the count as it starts from 0
            for i in range(1, row_number):
                old_file_name = worksheet_1.cell(row=i, column=1).value
                file_extension = worksheet_1.cell(row=i, column=2).value
                new_file_name = worksheet_1.cell(row=i, column=3).value
                old_file_name_plus_extension = old_file_name + "." + file_extension
                new_file_name_plus_extension = new_file_name + "." + file_extension
                main_file_list[old_file_name_plus_extension] = new_file_name_plus_extension
            self.listWidget.clear()
            self.listWidget.addItem("Displaying Values:\n")
            for key, values in main_file_list.items():
                text_to_print = f'{key} =====> {values}'
                self.listWidget.addItem(text_to_print)
        except Exception as e:
            self.listWidget.clear()
            self.listWidget.addItem("Something has gone wrong - see exception below:\n\n")
            self.listWidget.addItem(e)

    def abouttheauthor(self):
        open_author_window = About_The_Author()
        open_author_window.show()


class About_The_Author(QWidget):
    def __init__(self):
        super().__init__()
        self.setupUi(MainWindow)

    def setupUi(self, MainWindow):
        MainWindow.setObjectName("MainWindow")
        MainWindow.resize(541, 388)
        self.centralwidget = QtWidgets.QWidget(MainWindow)
        self.centralwidget.setObjectName("centralwidget")
        self.label = QtWidgets.QLabel(self.centralwidget)
        self.label.setGeometry(QtCore.QRect(160, 140, 271, 131))
        font = QtGui.QFont()
        font.setPointSize(16)
        self.label.setFont(font)
        self.label.setObjectName("label")
        MainWindow.setCentralWidget(self.centralwidget)
        self.menubar = QtWidgets.QMenuBar(MainWindow)
        self.menubar.setGeometry(QtCore.QRect(0, 0, 541, 21))
        self.menubar.setObjectName("menubar")
        MainWindow.setMenuBar(self.menubar)
        self.statusbar = QtWidgets.QStatusBar(MainWindow)
        self.statusbar.setObjectName("statusbar")
        MainWindow.setStatusBar(self.statusbar)
        self.retranslateUi(MainWindow)
        QtCore.QMetaObject.connectSlotsByName(MainWindow)

    def retranslateUi(self, MainWindow):
        _translate = QtCore.QCoreApplication.translate
        MainWindow.setWindowTitle(_translate("MainWindow", "MainWindow"))
        self.label.setText(_translate("MainWindow", "My name is craig"))


if __name__ == "__main__":
    import sys
    app = QtWidgets.QApplication(sys.argv)
    MainWindow = QtWidgets.QMainWindow()
    ui = Ui_MainWindow()
    ui.setupUi(MainWindow)
    MainWindow.show()
    time.sleep(0.1)
    sys.exit(app.exec_())
### END ###
